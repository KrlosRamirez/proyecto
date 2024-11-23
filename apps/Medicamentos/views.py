
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.forms import inlineformset_factory
import datetime
from datetime import datetime, timedelta
from django.utils import timezone
from .models import Benefactor, Medicamento, Destinatario, Entrega, EntregaItem, Insumo, Bitacora, RegistroMedicamento
from .forms import (
    MedicamentoForm,
    RegistroForm,
    CustomAuthenticationForm,
    DestinatarioForm,
    EntregaForm,
    EntregaItemForm,
    BaseEntregaItemFormSet,
    InsumoForm,
    Bitacora,
    BenefactorForm,
    RegistroMedicamentoForm
)
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from django.contrib.staticfiles import finders
from openpyxl.drawing.image import Image
from reportlab.pdfgen import canvas
from django.template.loader import get_template
from django.http import HttpResponse
from xhtml2pdf import pisa
from django.db import transaction
from django.core.paginator import Paginator
from django.db.models import Q
from django.shortcuts import redirect
from django.utils.timezone import now
import csv
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from .models import RegistroMedicamento
import tempfile
from django.conf import settings
from ..Utils.mensajes import *
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

function = Function()


#Incio de la app logueada.



def Inicio(request):
    return render(request, 'panel.html', {
    })




#-------------------------------------------------------------------------------------#




@login_required
@user_passes_test(lambda u: u.is_staff)
def bitacora(request):
    destinatario_id = request.GET.get('destinatario')
    if destinatario_id:
        registros = Bitacora.objects.filter(destinatario__id=destinatario_id).order_by('-fecha')
    else:
        registros = Bitacora.objects.all().order_by('-fecha')
    destinatarios = Destinatario.objects.all()
    return render(request, 'medicamentos/bitacora.html', {
        'registros': registros,
        'destinatarios': destinatarios,
        'destinatario_id': destinatario_id,
    })
def agregar_entrega(request):
    if request.method == 'POST':
        form = EntregaForm(request.POST)
        items_formset = EntregaItemForm(request.POST)
        if form.is_valid() and items_formset.is_valid():
            entrega = form.save()
            items_formset.instance = entrega
            items_formset.save()
            messages.success(request, 'Entrega registrada exitosamente.')
            # Registrar en la Bitácora
            Bitacora.objects.create(
                usuario=request.user,
                accion='Entrega registrada',
                descripcion=f'Se registró una entrega a {entrega.destinatario.nombre} {entrega.destinatario.apellido}.',
                destinatario=entrega.destinatario
            )
            return redirect('entregas')
    else:
        form = EntregaForm()
        items_formset = EntregaItemForm()
    return render(request, 'medicamentos/entrega_form.html', {'form': form, 'items_formset': items_formset})

@login_required
def exportar_bitacora_pdf(request):
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="bitacora.pdf"'
    p = canvas.Canvas(response)

    # Agregar contenido al PDF (puedes personalizarlo)
    p.drawString(100, 750, "Bitácora de Medicamentos")
    # Lógica adicional para extraer y mostrar registros en el PDF

    p.showPage()
    p.save()
    return response

@login_required
@user_passes_test(lambda u: u.is_staff)
def lista_insumos(request):
    insumos = Insumo.objects.all()
    return render(request, 'medicamentos/insumos_list.html', {'insumos': insumos})


@login_required
@user_passes_test(lambda u: u.is_staff)
def agregar_insumo(request):
    if request.method == 'POST':
        form = InsumoForm(request.POST)
        if form.is_valid():
            insumo = form.save()  # Código generado automáticamente en el modelo
            Bitacora.objects.create(
                usuario=request.user,
                accion="Agregar Insumo",
                descripcion=f"Se agregó el insumo {insumo.descripcion} con código {insumo.codigo}."
            )
            messages.success(request, "Insumo registrado exitosamente.")
            return redirect('medicamentos:lista_insumos')
    else:
        form = InsumoForm()
    return render(request, 'medicamentos/insumo_form.html', {'form': form})


@login_required
@user_passes_test(lambda u: u.is_staff)
def editar_insumo(request, pk):
    insumo = get_object_or_404(Insumo, pk=pk)
    if request.method == 'POST':
        form = InsumoForm(request.POST, instance=insumo)
        if form.is_valid():
            insumo = form.save()
            # Registrar en la Bitácora
            Bitacora.objects.create(
                usuario=request.user,
                accion="Editar Insumo",
                descripcion=f"Se editó el insumo {insumo.descripcion} con código {insumo.codigo}."
            )
            texto = function.mensaje("Insumo","Registro modificado exitosamente","success")
            messages.add_message(request, messages.SUCCESS,texto)
            return redirect('medicamentos:lista_insumos')
    else:
        form = InsumoForm(instance=insumo)
    return render(request, 'medicamentos/insumo_form.html', {'form': form})


def login_view(request):
    if request.user.is_authenticated:
        return redirect('index')
    if request.method == 'POST':
        form = CustomAuthenticationForm(request, data=request.POST)
        if form.is_valid():
            cedula = form.cleaned_data.get('username')
            password = form.cleaned_data.get('password')
            user = authenticate(username=cedula, password=password)
            if user is not None:
                login(request, user)
                messages.info(request, f'Has iniciado sesión como {cedula}.')
                return redirect('index')
            else:
                messages.error(request, 'Cédula de Identidad o contraseña incorrectos.')
        else:
            messages.error(request, 'Cédula de Identidad o contraseña incorrectos.')
    else:
        form = CustomAuthenticationForm()
    return render(request, 'medicamentos/login.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.info(request, 'Has cerrado sesión exitosamente.')
    return redirect('login')

def registro(request):
    if request.user.is_authenticated:
        return redirect('index')
    if request.method == 'POST':
        form = RegistroForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = False  # Cambia a True si deseas que el usuario sea administrador
            user.is_active = True
            user.save()
            cedula = form.cleaned_data.get('username')
            messages.success(request, f'Cuenta creada para {cedula}.')
            return redirect('login')
        else:
            messages.error(request, 'Por favor, corrige los errores a continuación.')
    else:
        form = RegistroForm()
    return render(request, 'medicamentos/registro.html', {'form': form})

@login_required
def index(request):
    return render(request, 'medicamentos/index.html')

@login_required
def lista_medicamentos(request):
    medicamentos = Medicamento.objects.all()

    registro_medicamentos = RegistroMedicamento.objects.all()
    # Fecha actual y fecha dentro de 30 días
    today = timezone.now().date()
    today_plus_30 = today + timedelta(days=30)
    # Medicamentos con stock bajo
    low_stock_medicamentos = registro_medicamentos.filter(cantidad__lte=10)
    # Medicamentos próximos a vencer
    about_to_expire_medicamentos = registro_medicamentos.filter(vencimiento_tanda__lte=today_plus_30)
    return render(request, 'medicamentos/medicamentos_list.html', {
        'registro_medicamentos': registro_medicamentos,
        'low_stock_medicamentos': low_stock_medicamentos,
        'about_to_expire_medicamentos': about_to_expire_medicamentos,
        'today_plus_30': today_plus_30,
        'today': today,
        'medicamentos': medicamentos
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def agregar_medicamento(request):
    if request.method == 'POST':
        form = MedicamentoForm(request.POST)
        if form.is_valid():
            medicamento = form.save()  

            Bitacora.objects.create(
                usuario=request.user,
                accion="Agregar Medicamento",
                descripcion=f"Se agregó el medicamento {medicamento.nombre} con código {medicamento.codigo}."
            )
            texto = function.mensaje("Medicamento", "Medicamento registrado exitosamente", "success")
            messages.add_message(request, messages.SUCCESS, texto)
            return redirect('medicamentos:lista_medicamentos')
    else:
        form = MedicamentoForm()
    return render(request, 'medicamentos/medicamento_form.html', {'form': form})

@login_required
def lista_regis_medicamentos(request):
    registro_medicamentos = RegistroMedicamento.objects.all()

    # Filtros para medicamentos en bajo stock y próximos a vencer
    low_stock_medicamentos = registro_medicamentos.filter(cantidad__lt=10)
    about_to_expire_medicamentos = registro_medicamentos.filter(
        vencimiento_tanda__lte=now().date() + timedelta(days=30),
        vencimiento_tanda__gte=now().date()
    )

    return render(request, 'medicamentos/medicamento_registro_list.html', {
        'registro_medicamentos': registro_medicamentos,
        'low_stock_medicamentos': low_stock_medicamentos,
        'about_to_expire_medicamentos': about_to_expire_medicamentos,
    })

@login_required
@user_passes_test(lambda u: u.is_staff)
def registrar_medicamento(request):
    if request.method == 'POST':
        form = RegistroMedicamentoForm(request.POST)

        if form.is_valid():
            registroMedicamento = form.save()

            codigo_medicamento = registroMedicamento.medicamento.pk
            medicamento = Medicamento.objects.get(pk=codigo_medicamento)
            medicamento.cantidad += registroMedicamento.cantidad
            medicamento.save()

            Bitacora.objects.create(
                usuario=request.user,
                accion="Registro Medicamento",
                descripcion=f"Se registro la entrada de {registroMedicamento.cantidad} unidades del medicamento {registroMedicamento.medicamento}."
            )
            texto = function.mensaje("Medicamento", "Medicamento registrado exitosamente", "success")
            messages.add_message(request, messages.SUCCESS, texto)
            return redirect('medicamentos:lista_medicamentos')
    else:
        form = RegistroMedicamentoForm()
    return render(request, 'medicamentos/medicamento_registro_form.html', {'form': form})


            



@login_required
@user_passes_test(lambda u: u.is_staff)
def editar_medicamento(request, pk):
    medicamento = get_object_or_404(Medicamento, pk=pk)
    if request.method == "POST":
        form = MedicamentoForm(request.POST, instance=medicamento)
        if form.is_valid():
            form.save() 
            texto = function.mensaje("Medicamento","Registro modificado exitosamente","success")
            messages.add_message(request, messages.SUCCESS,texto)
            return redirect('medicamentos:lista_medicamentos')
    else:
        form = MedicamentoForm(instance=medicamento)
    return render(request, 'medicamentos/medicamento_form.html', {"form":form,"medicamento":medicamento})


@login_required
@user_passes_test(lambda u: u.is_staff)
def eliminar_medicamento(request, pk):
    medicamento = get_object_or_404(Medicamento, pk=pk)
    if request.method == 'POST':
        medicamento.delete()
        texto = function.mensaje("Medicamento","Registro eliminado exitosamente","success")
        messages.add_message(request, messages.SUCCESS,texto)
        return redirect('medicamentos:lista_medicamentos')
    return render(request, 'medicamentos/eliminar_medicamento.html', {'medicamento': medicamento})


@login_required
@user_passes_test(lambda u: u.is_staff)
def editar_registro_medicamento(request, pk):
    medicamento = get_object_or_404(RegistroMedicamento, pk=pk)
    if request.method == "POST":
        form = RegistroMedicamentoForm(request.POST, instance=medicamento)
        if form.is_valid():
            form.save() 
            texto = function.mensaje("Medicamento","Registro modificado exitosamente","success")
            messages.add_message(request, messages.SUCCESS,texto)
            return redirect('medicamentos:lista_medicamentos')
    else:
        form = RegistroMedicamentoForm(instance=medicamento)
    return render(request, 'medicamentos/medicamento_registro_form.html', {"form":form,"medicamento":medicamento})


@login_required
@user_passes_test(lambda u: u.is_staff)
def eliminar_registro_medicamento(request, pk):
    medicamento = get_object_or_404(RegistroMedicamento, pk=pk)
    if request.method == 'POST':
        medicamento.delete()
        texto = function.mensaje("Medicamento","Registro eliminado exitosamente","success")
        messages.add_message(request, messages.SUCCESS,texto)
        return redirect('medicamentos:lista_medicamentos')
    return render(request, 'medicamentos/eliminar_registro_medicamento.html', {'medicamento': medicamento})




@login_required
def historial(request):
    entregas = Entrega.objects.all()
    return render(request, 'medicamentos/historial.html', {'entregas': entregas})

@login_required
def entregas(request):
    search_query = request.GET.get('search', '')
    if search_query:
        entregas_list = Entrega.objects.filter(
            Q(codigo__icontains=search_query) |
            Q(destinatario__nombre__icontains=search_query) |
            Q(destinatario__apellido__icontains=search_query)
        )
    else:
        entregas_list = Entrega.objects.all()
    entregas_list = entregas_list.order_by('-fecha_entrega')
    paginator = Paginator(entregas_list, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'medicamentos/entregas_list.html', {
        'entregas': page_obj,
        'is_paginated': True,
        'page_obj': page_obj,
        'paginator': paginator,
    })

@login_required
def detalle_entrega(request, pk):
    entrega = get_object_or_404(Entrega, pk=pk)
    return render(request, 'medicamentos/detalle_entrega.html', {'entrega': entrega})

@login_required
@user_passes_test(lambda u: u.is_staff)
def agregar_entrega(request):
    EntregaItemFormSet = inlineformset_factory(
        Entrega,
        EntregaItem,
        form=EntregaItemForm,
        formset=BaseEntregaItemFormSet,
        extra=1,
        can_delete=True,
    )
    if request.method == 'POST':
        form = EntregaForm(request.POST)
        if form.is_valid():
            entrega = form.save(commit=False)
            formset = EntregaItemFormSet(request.POST, instance=entrega)
            if formset.is_valid():
                # Validar cantidades antes de guardar
                for form_item in formset:
                    medicamento = form_item.cleaned_data.get('medicamento')
                    benefactor_entrega = form_item.cleaned_data.get('benefactor')
                    cantidad = form_item.cleaned_data.get('cantidad')
                    lote = form_item.cleaned_data.get('lote')
                    cantidad_en_inventario = RegistroMedicamento.objects.get(pk=lote.id)

                    if cantidad > cantidad_en_inventario.cantidad:
                        messages.error(
                            request,
                            f"No hay suficiente stock de {medicamento.nombre} en el lote {lote.identificador}. Disponible: {cantidad_en_inventario.cantidad}"
                        )
                        return render(request, 'medicamentos/entrega_form.html', {'form': form, 'formset': formset})

                # Guardar la entrega y los items
                entrega.save()
                formset.save()

                # Actualizar inventario dentro del bucle
                for form_item in formset:
                    lote = form_item.cleaned_data.get('lote')
                    cantidad = form_item.cleaned_data.get('cantidad')
                    cantidad_en_inventario = RegistroMedicamento.objects.get(pk=lote.id)
                    cantidad_en_inventario.cantidad -= cantidad

                    if cantidad_en_inventario.cantidad > 0:
                        cantidad_en_inventario.save()
                    else:
                        cantidad_en_inventario.delete()

                # Actualizar stock total del medicamento
                for item in entrega.items.all():
                    medicamento = item.medicamento
                    medicamento.cantidad -= item.cantidad
                    medicamento.save()

                messages.success(request, 'Entrega registrada exitosamente.')
                return redirect('medicamentos:entregas')
            else:
                messages.error(request, 'Por favor, corrige los errores en los medicamentos.')
        else:
            messages.error(request, 'Por favor, corrige los errores en la información de la entrega.')
            formset = EntregaItemFormSet(request.POST)
    else:
        form = EntregaForm()
        entrega = Entrega()
        formset = EntregaItemFormSet(instance=entrega)
    return render(request, 'medicamentos/entrega_form.html', {'form': form, 'formset': formset})


@login_required
@user_passes_test(lambda u: u.is_staff)
def agregar_destinatario(request):
    if request.method == 'POST':
        form = DestinatarioForm(request.POST)
        if form.is_valid():
            form.save()
            texto = function.mensaje("Benefactor","Benefactor agrego exitosamente","success")
            messages.add_message(request, messages.SUCCESS,texto)
            return redirect('medicamentos:lista_destinatarios')
        else:
            messages.error(request, 'Por favor, corrige los errores a continuación.')
    else:
        form = DestinatarioForm()
    return render(request, 'medicamentos/destinatario_form.html', {'form': form})

@login_required
@user_passes_test(lambda u: u.is_staff)
def editar_destinatario(request, pk):
    destinatario = get_object_or_404(Destinatario, pk=pk)
    if request.method == 'POST':
        form = DestinatarioForm(request.POST, instance=destinatario)
        if form.is_valid():
            form.save()
            texto = function.mensaje("Benefactor","modificacion exitosamente","success")
            messages.add_message(request, messages.SUCCESS,texto)
            return redirect('medicamentos:lista_destinatarios')
        else:
            messages.error(request, 'Por favor, corrige los errores a continuación.')
    else:
        form = DestinatarioForm(instance=destinatario)
    return render(request, 'medicamentos/destinatario_form.html', {'form': form})

@login_required
def detalle_destinatario(request, pk):
    destinatario = get_object_or_404(Destinatario, pk=pk)
    return render(request, 'medicamentos/detalle_destinatario.html', {'destinatario': destinatario})

@login_required
def lista_destinatarios(request):
    destinatarios = Destinatario.objects.all()
    return render(request, 'medicamentos/destinatarios_list.html', {'destinatarios': destinatarios})

@login_required
@user_passes_test(lambda u: u.is_staff)
def eliminar_destinatario(request, pk):
    destinatario = get_object_or_404(Destinatario, pk=pk)
    if request.method == 'POST':
        destinatario.delete()
        texto = function.mensaje("Destinatario","Registro eliminado exitosamente","success")
        messages.add_message(request, messages.SUCCESS,texto)
        return redirect('medicamentos:lista_destinatarios')    
    return render(request, 'medicamentos/eliminar_destinatario.html', {'destinatario': destinatario})


@login_required
def exportar_entregas_csv(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="entregas.csv"'

    writer = csv.writer(response)
    writer.writerow(['Código', 'Destinatario', 'Medicamento', 'Cantidad', 'Fecha de Entrega', 'Observaciones'])

    entregas = Entrega.objects.all()
    for entrega in entregas:
        for item in entrega.items.all():
            writer.writerow([
                entrega.codigo,
                entrega.destinatario.nombre,
                item.medicamento.nombre,
                item.cantidad,
                entrega.fecha_entrega,
                entrega.observaciones,
            ])


    return response

#####################################################################################################

@login_required
@user_passes_test(lambda u: u.is_staff)
def agregar_benefactor(request):
    if request.method == 'POST':
        form = BenefactorForm(request.POST)
        if form.is_valid():
            form.save()
            texto = "Benefactor registrado exitosamente."
            messages.success(request, texto)
            return redirect('medicamentos:lista_benefactores')
        else:
            messages.error(request, 'Por favor, corrige los errores a continuación.')
    else:
        form = BenefactorForm()

    return render(request, 'medicamentos/benefactor_form.html', {'form': form})


@login_required
@user_passes_test(lambda u: u.is_staff)
def editar_benefactor(request, pk):
    benefactor = get_object_or_404(Benefactor, pk=pk)
    if request.method == 'POST':
        form = BenefactorForm(request.POST, instance=benefactor)
        if form.is_valid():
            form.save()
            texto = "Datos modificado Realizados"
            messages.success(request, texto)
            return redirect('medicamentos:lista_benefactores')
        else:
            messages.error(request, 'Por favor, corrige los errores a continuación.')
    else:
        form = BenefactorForm(instance=benefactor)

    return render(request, 'medicamentos/benefactor_form.html', {'form': form})


@login_required
def detalle_benefactor(request, pk):
    benefactores = get_object_or_404(Benefactor, pk=pk)
    return render(request, 'medicamentos/detalle_benefactor.html', {'benefactores': benefactores})

@login_required
def lista_benefactores(request):  
    benefactores = Benefactor.objects.all()
    return render(request, 'medicamentos/benefactores_list.html', {'benefactores': benefactores})

@login_required
@user_passes_test(lambda u: u.is_staff)
def eliminar_benefactor(request, pk):
    benefactor = get_object_or_404(Benefactor, pk=pk)
    if request.method == 'POST':
        benefactor.delete()
        texto = "Registro eliminado exitosamente"
        messages.success(request, texto)
        return redirect('medicamentos:lista_benefactores')

    return render(request, 'medicamentos/eliminar_benefactor.html', {'benefactor': benefactor})

@login_required
@user_passes_test(lambda u: u.is_staff)
def reporte_registro_medicamentos(request):
    registros = RegistroMedicamento.objects.all()

    template_path = 'medicamentos/reporte_registro_medicamentos.html'
    context = {'registros': registros}

    # Crear una respuesta HTTP con el tipo de contenido PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="reporte_registro_medicamentos.pdf"'

    # Encontrar el template y renderizarlo
    template = get_template(template_path)
    html = template.render(context)

    # Crear el PDF
    pisa_status = pisa.CreatePDF(
       html, dest=response, link_callback=link_callback
    )

    # Si hay un error, mostrarlo
    if pisa_status.err:
       return HttpResponse('Ocurrió un error al generar el PDF: %s' % pisa_status.err)
    return response

def link_callback(uri, rel):
    # Resolver la ruta absoluta de los archivos estáticos y media
    sUrl = settings.STATIC_URL      # /static/
    sRoot = settings.STATIC_ROOT    # /var/www/example.com/static/
    mUrl = settings.MEDIA_URL       # /media/
    mRoot = settings.MEDIA_ROOT     # /var/www/example.com/media/

    if uri.startswith(mUrl):
        path = os.path.join(mRoot, uri.replace(mUrl, ""))
    elif uri.startswith(sUrl):
        path = os.path.join(sRoot, uri.replace(sUrl, ""))
    else:
        return uri  # Devolver la URI sin cambios si no es estático ni media

    # Asegurarse de que el archivo existe
    if not os.path.isfile(path):
        raise Exception('El archivo no existe: %s' % path)
    return path


@login_required
@user_passes_test(lambda u: u.is_staff)
def generar_reporte_entregas(request):
    # Crear un nuevo libro de Excel y seleccionar la hoja activa
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Listado de Entregas"

    # Insertar filas para dejar espacio para el logo
    sheet.insert_rows(idx=1, amount=5)  # 5 filas para el logo

    # Definir estilos
    titulo_font = Font(size=14, bold=True)
    header_font = Font(color='FFFFFFFF', bold=True)  # Letras blancas
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='00008080', end_color='00008080', fill_type='solid')

    # Añadir título al reporte
    sheet.merge_cells('A6:E6')  # La fila 6 después de las 5 filas del logo
    sheet['A6'] = 'Reporte de Entregas de Medicamentos Caritas'
    sheet['A6'].font = titulo_font
    sheet['A6'].alignment = center_alignment

    # Añadir encabezados en la fila 7
    encabezados = ['Código', 'Destinatario', 'Fecha de Entrega', 'Medicamentos', 'Observaciones']
    for col_num, header in enumerate(encabezados, 1):
        cell = sheet.cell(row=7, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # Ajustar el ancho de las columnas
    column_widths = [10, 25, 18, 40, 30]
    for i, column_width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

     # Encontrar la ruta del logo
    logo_path = finders.find('assets/logo/logocaritas.png')
    if logo_path:  # Si el logo existe, agregarlo
        logo = Image(logo_path)
        logo.width = 120  # Ajusta el tamaño del logo según sea necesario
        logo.height = 120
        sheet.add_image(logo, 'A1')

    # Agregar los datos de las entregas, comenzando en la fila 8
    entregas = Entrega.objects.all()
    row_num = 8  # Inicia en la fila 8
    for entrega in entregas:
        medicamentos = ", ".join([f"{item.medicamento.nombre} ({item.cantidad})" for item in entrega.items.all()])
        data = [
            entrega.pk,
            f"{entrega.destinatario.nombre} {entrega.destinatario.apellido}",
            entrega.fecha_entrega.strftime('%d/%m/%Y'),
            medicamentos,
            entrega.observaciones or ''
        ]
        for col_num, value in enumerate(data, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
        row_num += 1

    # Aplicar bordes y alineación a todas las celdas con datos
    for row in sheet.iter_rows(min_row=8, min_col=1, max_col=5, max_row=row_num - 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Preparar el archivo para descarga
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Reporte_Entregas.xlsx"'
    workbook.save(response)
    return response


def generar_reporte_registro_medicamentos(request):
    # Crear un nuevo libro de Excel y seleccionar la hoja activa
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Registro Medicamentos"

    # Insertar filas para dejar espacio para el logo
    sheet.insert_rows(idx=1, amount=5)

    # Encontrar la ruta del logo
    logo_path = finders.find('assets/logo/logocaritas.png')
    if logo_path:  # Verificar si el logo fue encontrado
        logo = Image(logo_path)
        logo.width = 120  # Ajustar tamaño del logo
        logo.height = 120
        sheet.add_image(logo, 'A1')  # Colocar el logo en la celda A1

    # Definir estilos
    titulo_font = Font(size=14, bold=True)
    header_font = Font(color='FFFFFFFF', bold=True)  # Letras blancas
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='00008080', end_color='00008080', fill_type='solid')

    # Añadir título al reporte
    sheet.merge_cells('A6:H6')
    sheet['A6'] = 'Reporte de Registro de Medicamentos'
    sheet['A6'].font = titulo_font
    sheet['A6'].alignment = center_alignment

    # Añadir encabezados en la fila 7
    encabezados = ['Código de Registro', 'Medicamento', 'Cantidad', 'Fecha de Registro', 'Fecha de Producción', 'Fecha de Vencimiento', 'Benefactor']
    for col_num, header in enumerate(encabezados, 1):
        cell = sheet.cell(row=7, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # Ajustar el ancho de las columnas
    column_widths = [20, 30, 10, 20, 20, 20, 30]
    for i, column_width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

    # Agregar los datos de los registros
    registros = RegistroMedicamento.objects.all()
    row_num = 8  # Inicia en la fila 8
    for registro in registros:
        data = [
            registro.pk,
            registro.medicamento.nombre if registro.medicamento else "Sin Medicamento",
            registro.cantidad,
            registro.fecha_registro_medicamento.strftime('%d/%m/%Y'),
            registro.produccion_tanda.strftime('%d/%m/%Y'),
            registro.vencimiento_tanda.strftime('%d/%m/%Y'),
            registro.benefactor.nombre if registro.benefactor else "Sin Benefactor"
        ]
        for col_num, value in enumerate(data, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
        row_num += 1

    # Aplicar bordes y alineación a todas las celdas con datos
    for row in sheet.iter_rows(min_row=8, min_col=1, max_col=7, max_row=row_num - 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Preparar el archivo para descarga
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Reporte_Registro_Medicamentos.xlsx"'
    workbook.save(response)
    return response

def reporte_bajo_stock(request):
    # Crear un nuevo libro de Excel y seleccionar la hoja activa
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Medicamentos Bajo Stock"

    # Insertar filas para dejar espacio para el logo
    sheet.insert_rows(idx=1, amount=5)

    # Encontrar la ruta del logo
    logo_path = finders.find('assets/logo/logocaritas.png')
    if logo_path:
        logo = Image(logo_path)
        logo.width = 120
        logo.height = 120
        sheet.add_image(logo, 'A1')

    # Definir estilos
    titulo_font = Font(size=14, bold=True)
    header_font = Font(color='FFFFFFFF', bold=True)  # Letras blancas
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='00008080', end_color='00008080', fill_type='solid')

    # Añadir título al reporte
    sheet.merge_cells('A6:E6')
    sheet['A6'] = 'Reporte de Medicamentos con Bajo Stock'
    sheet['A6'].font = titulo_font
    sheet['A6'].alignment = center_alignment

    # Añadir encabezados en la fila 7
    encabezados = ['Código', 'Medicamento', 'Cantidad', 'Fecha de Registro', 'Benefactor']
    for col_num, header in enumerate(encabezados, 1):
        cell = sheet.cell(row=7, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # Ajustar el ancho de las columnas
    column_widths = [10, 30, 10, 20, 30]
    for i, column_width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

    # Obtener registros con bajo stock (cantidad < 10)
    registros = RegistroMedicamento.objects.filter(cantidad__lt=10)
    row_num = 8  # Inicia en la fila 8
    for registro in registros:
        data = [
            registro.pk,
            registro.medicamento.nombre if registro.medicamento else "Sin Medicamento",
            registro.cantidad,
            registro.fecha_registro_medicamento.strftime('%d/%m/%Y'),
            registro.benefactor.nombre if registro.benefactor else "Sin Benefactor"
        ]
        for col_num, value in enumerate(data, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
        row_num += 1

    # Aplicar bordes y alineación a todas las celdas con datos
    for row in sheet.iter_rows(min_row=8, min_col=1, max_col=5, max_row=row_num - 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Preparar el archivo para descarga
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Reporte_Bajo_Stock.xlsx"'
    workbook.save(response)
    return response


def reporte_proximos_a_vencer(request):
    # Crear un nuevo libro de Excel y seleccionar la hoja activa
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Medicamentos Próximos a Vencer"

    # Insertar filas para dejar espacio para el logo
    sheet.insert_rows(idx=1, amount=5)

    # Encontrar la ruta del logo
    logo_path = finders.find('assets/logo/logocaritas.png')
    if logo_path:
        logo = Image(logo_path)
        logo.width = 120
        logo.height = 120
        sheet.add_image(logo, 'A1')

    # Definir estilos
    titulo_font = Font(size=14, bold=True)
    header_font = Font(color='FFFFFFFF', bold=True)
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='00008080', end_color='00008080', fill_type='solid')

    # Añadir título al reporte
    sheet.merge_cells('A6:F6')
    sheet['A6'] = 'Reporte de Medicamentos Próximos a Vencer'
    sheet['A6'].font = titulo_font
    sheet['A6'].alignment = center_alignment

    # Añadir encabezados en la fila 7
    encabezados = ['Código', 'Medicamento', 'Fecha de Vencimiento', 'Días Restantes', 'Cantidad', 'Benefactor']
    for col_num, header in enumerate(encabezados, 1):
        cell = sheet.cell(row=7, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
        cell.fill = header_fill

    # Ajustar el ancho de las columnas
    column_widths = [10, 30, 20, 15, 10, 30]
    for i, column_width in enumerate(column_widths, 1):
        sheet.column_dimensions[openpyxl.utils.get_column_letter(i)].width = column_width

    # Obtener la fecha actual y calcular el rango de próximos 30 días
    fecha_actual = timezone.now().date()
    fecha_limite = fecha_actual + timedelta(days=30)

    # Obtener registros que vencen en los próximos 30 días
    registros = RegistroMedicamento.objects.filter(
        vencimiento_tanda__lte=fecha_limite,
        vencimiento_tanda__gte=fecha_actual
    )

    row_num = 8  # Inicia en la fila 8
    for registro in registros:
        dias_restantes = (registro.vencimiento_tanda - fecha_actual).days
        data = [
            registro.pk,
            registro.medicamento.nombre if registro.medicamento else "Sin Medicamento",
            registro.vencimiento_tanda.strftime('%d/%m/%Y'),
            dias_restantes,
            registro.cantidad,
            registro.benefactor.nombre if registro.benefactor else "Sin Benefactor"
        ]
        for col_num, value in enumerate(data, 1):
            cell = sheet.cell(row=row_num, column=col_num, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border
        row_num += 1

    # Aplicar bordes y alineación
    for row in sheet.iter_rows(min_row=8, min_col=1, max_col=6, max_row=row_num - 1):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Preparar para descarga
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="Reporte_Proximos_A_Vencer.xlsx"'
    workbook.save(response)
    return response